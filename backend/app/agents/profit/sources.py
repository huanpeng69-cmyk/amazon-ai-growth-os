"""利润模块数据来源（Data Sources）。

满足两种真实数据接入方式（均非编造）：
  1. manual         —— 用户在表单中手动输入的成本（直接透传）
  2. excel / csv    —— 上传成本表，解析为结构化成本字段

平台费用（FBA 履约费 / 佣金率）由统一数据层 amazon_connector 的真实费率表提供
（见 app.data.dal.get_fee_schedule），不再使用任何 Mock 供应链连接器。

解析出的字段仅作为 Profit Agent 的输入，所有金额均为用户/接口真实提供，
不经由 LLM 编造。
"""
from __future__ import annotations

import csv
import io
import re
from typing import Dict, Optional

# 成本字段中文/英文同义词 → 规范字段名
_SYNONYMS = {
    "selling_price": ["售价", "价格", "卖价", "selling price", "price", "list price"],
    "product_cost": ["产品成本", "采购成本", "制造成本", "成本价", "product cost", "cost", "unit cost"],
    "shipping_cost": ["头程", "物流", "运费", "头程物流", "shipping", "freight", "logistics"],
    "fba_fee": ["fba", "fba费", "履约费", "配送费", "fulfillment fee", "fba fee"],
    "referral_fee_rate": ["佣金", "佣金率", "平台佣金", "referral", "commission", "referral fee"],
    "ad_acos": ["广告", "acos", "广告费", "广告花费", "ad acos", "ad spend", "acos"],
    "other_cost_per_unit": ["其他", "包装", "其他成本", "杂费", "other", "packaging", "misc"],
    "monthly_fixed_cost": ["月固定", "固定成本", "月度固定", "monthly fixed", "fixed cost", "overhead"],
    "initial_investment": ["首单", "投入", "首单投入", "备货", "investment", "initial", "capex"],
}

_RATE_FIELDS = {"referral_fee_rate", "ad_acos"}


def _norm(s: str) -> str:
    return re.sub(r"\s+", "", str(s).lower())


def _match_field(label: str) -> Optional[str]:
    n = _norm(label)
    if not n:
        return None
    for field, syns in _SYNONYMS.items():
        for s in syns:
            if _norm(s) == n or _norm(s) in n or n in _norm(s):
                return field
    return None


def _to_number(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("%", "").replace("$", "").strip()
    if s == "":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def extract_cost_fields(rows: list[list]) -> Dict[str, float]:
    """从二维单元格中识别「标签-数值」对，映射为规范成本字段。"""
    found: Dict[str, float] = {}

    # 形态 A：第一行是表头（字符串），第二行是对应的数值
    if len(rows) >= 2:
        header, data = rows[0], rows[1]
        if all(isinstance(h, str) and h.strip() for h in header):
            for i, h in enumerate(header):
                fld = _match_field(h)
                if fld and i < len(data):
                    num = _to_number(data[i])
                    if num is not None and fld not in found:
                        found[fld] = num

    # 形态 B：逐行「名称 | 数值」
    for row in rows:
        if not row:
            continue
        if len(row) >= 2:
            fld = _match_field(row[0])
            if fld:
                num = _to_number(row[1])
                if num is not None and fld not in found:
                    found[fld] = num

    # 比率字段：若以百分比整数/小数形式给出（>1），归一化为 0-1
    for fld in _RATE_FIELDS:
        if fld in found and found[fld] > 1:
            found[fld] = round(found[fld] / 100.0, 4)
    return found


def parse_cost_bytes(raw: bytes, filename: str) -> Dict[str, float]:
    """解析上传的成本文件（.xlsx / .xls / .csv）。"""
    name = (filename or "").lower()
    if name.endswith((".csv", ".txt")):
        text = raw.decode("utf-8-sig", errors="replace")
        rows = [r for r in csv.reader(io.StringIO(text)) if any(c.strip() for c in r)]
        return extract_cost_fields(rows)

    # Excel：优先 openpyxl
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        raise RuntimeError("解析 .xlsx 需要 openpyxl，请先 pip install openpyxl，或改用 .csv 上传")
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows = [[c.value for c in row] for row in ws.iter_rows()]
    wb.close()
    return extract_cost_fields(rows)
